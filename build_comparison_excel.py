#!/usr/bin/env python3
"""Build Excel workbook from comparison_results.json."""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = Path("eval_final/comparison_results.json")
OUT = Path("eval_final/comparison_results.xlsx")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ALT_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="B4C6E7")

PARCEL_ORDER = ["001296507", "017028141", "027266834", "027266966"]


def auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, 50)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_parcel_sheet(wb, data):
    """One sheet with all 4 parcels × 2 pipelines × 2 boundary modes."""
    ws = wb.create_sheet(title="Parcel Comparison")
    headers = ["Parcel ID", "Pipeline", "Boundary Mode", "Risk Score",
               "Zone 1a Density", "Zone 1b Density", "Zone 2 Density",
               "Zone 3 Density", "Min Veg Distance (m)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    row = 2
    for pipeline, key, label in [("High-Res (Drone/CNN)", "highres_parcels", "highres"),
                                  ("Low-Res (Satellite/MS+Meta)", "lowres_parcels", "lowres")]:
        parcels = sorted(data[key], key=lambda x: x["pid"])
        for p in parcels:
            for mode_label, mode_key in [("Without Parcel", "without_parcel"),
                                          ("With Parcel", "with_parcel")]:
                d = p[mode_key]
                if d["risk_score"] is None:
                    continue
                vals = [
                    p["pid"], pipeline, mode_label,
                    d["risk_score"], d["zone_1a"], d["zone_1b"],
                    d["zone_2"], d["zone_3"], d["min_veg_dist"],
                ]
                for c, v in enumerate(vals, 1):
                    ws.cell(row=row, column=c, value=v)
                if row % 2 == 0:
                    for c2 in range(1, len(headers) + 1):
                        ws.cell(row=row, column=c2).fill = ALT_FILL
                row += 1
        # Blank row between pipelines
        row += 1

    style_header(ws, len(headers))
    auto_width(ws)


def build_per_parcel_sheets(wb, data):
    """One sheet per parcel with 4 rows (high/low × with/without)."""
    hr = {p["pid"]: p for p in data["highres_parcels"]}
    lr = {p["pid"]: p for p in data["lowres_parcels"]}

    for pid in PARCEL_ORDER:
        ws = wb.create_sheet(title=f"Parcel {pid}")
        headers = ["Pipeline", "Boundary Mode", "Risk Score",
                   "Zone 1a", "Zone 1b", "Zone 2", "Zone 3", "Min Veg Dist (m)"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)

        row = 2
        for pipeline, pdata in [("High-Res (Drone/CNN)", hr.get(pid)),
                                 ("Low-Res (Satellite/MS+Meta)", lr.get(pid))]:
            if pdata is None:
                continue
            for mode_label, mode_key in [("Without Parcel", "without_parcel"),
                                          ("With Parcel", "with_parcel")]:
                d = pdata[mode_key]
                if d["risk_score"] is None:
                    continue
                vals = [pipeline, mode_label, d["risk_score"],
                        d["zone_1a"], d["zone_1b"], d["zone_2"],
                        d["zone_3"], d["min_veg_dist"]]
                for c, v in enumerate(vals, 1):
                    ws.cell(row=row, column=c, value=v)
                if row % 2 == 0:
                    for c2 in range(1, len(headers) + 1):
                        ws.cell(row=row, column=c2).fill = ALT_FILL
                row += 1

        style_header(ws, len(headers))
        auto_width(ws)


def build_community_sheet(wb, data):
    """Community-level risk scores."""
    ws = wb.create_sheet(title="Community Scores")
    headers = ["Community", "Buildings", "With Vegetation", "Mean (All)",
               "Mean (Filtered)", "Median", "Max"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    entries = [
        ("Cumberland Drone (no parcel)", "cumberland_drone_no_parcel"),
        ("Cumberland Drone (w/ parcel)", "cumberland_drone_with_parcel"),
        ("Cumberland Sat (no parcel)", "cumberland_sat_no_parcel"),
        ("Cumberland Sat (w/ parcel)", "cumberland_sat_with_parcel"),
        ("Logan Lake (no parcel)", "logan_lake_no_parcel"),
        ("Logan Lake (w/ parcel)", "logan_lake_with_parcel"),
        ("West Kelowna (no parcel)", "west_kelowna_no_parcel"),
    ]

    row = 2
    for label, key in entries:
        d = data[key]
        vals = [label, d["total"], d["with_veg"], d["mean_all"],
                d["mean_filtered"], d["median"], d["max"]]
        for c, v in enumerate(vals, 1):
            ws.cell(row=row, column=c, value=v)
        if row % 2 == 0:
            for c2 in range(1, len(headers) + 1):
                ws.cell(row=row, column=c2).fill = ALT_FILL
        row += 1

    style_header(ws, len(headers))
    auto_width(ws)


def main():
    with open(DATA) as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_parcel_sheet(wb, data)
    build_per_parcel_sheets(wb, data)
    build_community_sheet(wb, data)

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
