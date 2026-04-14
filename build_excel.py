#!/usr/bin/env python3
"""Build a single Excel workbook from all eval_final results."""

import csv
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EVAL_DIR = Path("eval_final")
OUT = EVAL_DIR / "eval_results.xlsx"

REGIONS = ["cumberland", "west_kelowna", "logan_lake", "silver_star"]
NICE = {"cumberland": "Cumberland", "west_kelowna": "West Kelowna",
        "logan_lake": "Logan Lake", "silver_star": "Silver Star"}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
ALT_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(bottom=THIN)


def auto_width(ws):
    for col in ws.columns:
        mx = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                mx = max(mx, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(mx + 3, 50)


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def add_csv_sheet(wb, name, csv_path):
    """Add a sheet from a CSV file."""
    ws = wb.create_sheet(title=name[:31])  # Excel 31 char limit
    with open(csv_path) as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader, 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c)
                # Try to convert numeric
                try:
                    cell.value = float(val)
                    if cell.value == int(cell.value) and abs(cell.value) < 1e9:
                        cell.value = int(cell.value)
                except (ValueError, OverflowError):
                    cell.value = val
            if r > 1 and r % 2 == 0:
                for c2 in range(1, len(row) + 1):
                    ws.cell(row=r, column=c2).fill = ALT_FILL
    if ws.max_column:
        style_header(ws, ws.max_column)
    auto_width(ws)


def build_summary(wb):
    """Build a cross-region summary sheet."""
    ws = wb.create_sheet(title="Summary", index=0)

    headers = ["Region", "Type", "Buildings", "Mean Risk", "Median Risk",
               "High Risk %", "Moderate Risk %", "Low Risk %",
               "Non-Compliant %", "Partially Compliant %", "Compliant %",
               "Mean Veg Distance (m)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    row_idx = 2
    for region in REGIONS:
        for mode in ["tiles", "combined", "parcels"]:
            stats_path = EVAL_DIR / f"{region}_{mode}" / "community_statistics.csv"
            if not stats_path.exists():
                continue

            data = {}
            with open(stats_path) as f:
                for r in csv.DictReader(f):
                    data[(r["category"], r["metric"])] = r["value"]

            def g(cat, met):
                v = data.get((cat, met), "")
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return v

            vals = [
                NICE[region],
                mode.title(),
                g("overview", "total_buildings") or g("overview", "total_parcels") or g("overview", "parcels_with_buildings"),
                g("risk_score", "mean"),
                g("risk_score", "median"),
                g("risk_category", "high_percent"),
                g("risk_category", "moderate_percent"),
                g("risk_category", "low_percent"),
                g("compliance", "non_compliant_percent"),
                g("compliance", "partially_compliant_percent"),
                g("compliance", "compliant_percent"),
                g("vegetation_distance", "mean"),
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row_idx, column=c, value=v)
            if row_idx % 2 == 0:
                for c2 in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c2).fill = ALT_FILL
            row_idx += 1

    style_header(ws, len(headers))
    auto_width(ws)


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Add raw data sheets for each region/mode
    for region in REGIONS:
        for mode, csv_name in [("tiles", "building_risk_scores.csv"),
                                ("combined", "building_risk_scores.csv"),
                                ("parcels", "parcel_risk_scores.csv")]:
            csv_path = EVAL_DIR / f"{region}_{mode}" / csv_name
            if csv_path.exists():
                sheet_name = f"{NICE[region]} {mode.title()}"
                add_csv_sheet(wb, sheet_name, csv_path)

        # Add community stats sheets
        for mode in ["tiles", "combined", "parcels"]:
            stats_path = EVAL_DIR / f"{region}_{mode}" / "community_statistics.csv"
            if stats_path.exists():
                sheet_name = f"{NICE[region]} {mode.title()} Stats"
                add_csv_sheet(wb, sheet_name, stats_path)

    # Build summary as first sheet
    build_summary(wb)

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
